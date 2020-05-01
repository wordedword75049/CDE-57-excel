import argparse
import os
import pathlib
import random
import sys
import xlsxwriter
import csv
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import win32com.client as win32
from openpyxl.chart.layout import Layout, ManualLayout


def read_csv(rowNumber, filename):
    csv_path = filename
    extracted_data = []
    with open(csv_path, "r") as f_obj:
        reader = csv.reader(f_obj)
        for row in reader:
            extracted_data.append(row[rowNumber])
    return extracted_data

def generate_labels(size, textmode):
    labels = []
    random.seed()
    mode = random.randint(1, 4)
    if mode == 1:
        labels = list(range(1, size+1))
    elif mode == 2:
        data = read_csv(textmode, "res\constituents.csv")
        labels = random.sample(data[1:], size)
    elif mode == 3:
        start_year = random.randint(2010, 2020)
        labels = list(range(start_year-size, start_year))
    elif mode == 4:
        months = read_csv(textmode, "res\months.csv")
        start_year = random.randint(2010-(size // 12), 2020-(size // 12))
        start_month = random.randint(1,12)
        for i in range(size):
            if textmode == 0:
                labels.append(str(months[(start_month + i) % 12]) + "." + str(start_year + ((start_month + i) // 12)))
            elif textmode == 1:
                labels.append(str(months[(start_month + i) % 12]) + " " + str(start_year+((start_month + i) // 12)))

    return labels

def create_table(path, MaxColumnNumber, textmode):
    random.seed()
    ColumnCount = random.randint(5, MaxColumnNumber)
    full_path = str(path) + '\source.xlsx'
    workbook = xlsxwriter.Workbook(full_path )

    worksheet = workbook.add_worksheet()
    LowerBound = random.randint(100, 500)
    UpperBound = random.randint(600, 1000)
    x_label = generate_labels(ColumnCount, textmode)
    for column in range(ColumnCount):
        value = random.randint(LowerBound, UpperBound)
        worksheet.write(column, 0, x_label[column])
        worksheet.write(column, 1, value)
    workbook.close()

    OpenedWb = load_workbook(full_path )
    OpenedWS = OpenedWb['Sheet1']
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = random.randint(1, 8)
    chart1.title = ""
    chart1.y_axis.title = ''
    chart1.x_axis.title = ''

    data = Reference(OpenedWS, min_col=2, min_row=1, max_row=ColumnCount)
    cats = Reference(OpenedWS, min_col=1, min_row=1, max_row=ColumnCount)
    chart1.add_data(data, titles_from_data= True)
    chart1.set_categories(cats)
    chart1.shape = 10
    chart1.layout = Layout( #документация по параметрам layout диаграммы - https://openpyxl.readthedocs.io/en/stable/charts/chart_layout.html#size-and-position
        manualLayout=ManualLayout(
            h=0.85, w=0.85,
        )
    )
    OpenedWS.add_chart(chart1, "D3")
    OpenedWb.save(full_path)

def export_image(path):
    app = win32.Dispatch('Excel.Application')
    workbook_file_name = str(path) +  '\\source.xlsx'
    workbook = app.Workbooks.Open(Filename=workbook_file_name)
    app.DisplayAlerts = False

    for sheet in workbook.Worksheets:
        for chartObject in sheet.ChartObjects():
            chartObject.Activate()
            image_file_name=str(path) +  '\\image.png'
            chartObject.Chart.Export(image_file_name)
    workbook.Close(SaveChanges=False, Filename=workbook_file_name)


def generate(args):
    mode = 0
    if args.TextMode == "short":
        mode = 0
    elif args.TextMode == "long":
        mode = 1
    dirpath = os.path.join(args.BasePath, args.BaseName + '\\Data')
    if not os.access(str(dirpath), os.F_OK):
        os.makedirs(str(dirpath))
    for iteration in range(args.ChartsCount):
        iteration_path = os.path.join(dirpath, str(iteration+1))
        if not os.access(iteration_path, os.F_OK):
            os.makedirs(iteration_path)
        create_table(iteration_path, args.MaxColumnsCount, mode)
        export_image(iteration_path)



def parse(argv):
    parser = argparse.ArgumentParser(description="""
     Parser of image and tables generator
        """, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('BasePath', action='store', type=pathlib.Path)
    parser.add_argument('BaseName', action='store', type=str)
    parser.add_argument('ChartsCount', action='store', type=int)
    parser.add_argument('MaxColumnsCount', action='store', type=int)
    parser.add_argument('TextMode', action='store', type=str)

    return parser.parse_args(argv[1:])


def main():
    args = parse(sys.argv)
    generate(args)


main()
