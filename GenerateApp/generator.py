import argparse
import os
import pathlib
import random
import sys
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import win32com.client as win32

def create_table(path):
    random.seed()
    ColumnCount = random.randint(5, 12)
    full_path = str(path) + '\source.xlsx'
    workbook = xlsxwriter.Workbook(full_path )

    worksheet = workbook.add_worksheet()
    LowerBound = random.randint(100, 500)
    UpperBound = random.randint(600, 1000)
    for column in range(ColumnCount):
        value = random.randint(LowerBound, UpperBound)
        worksheet.write(column, 0, column + 1)
        worksheet.write(column, 1, value)
    workbook.close()

    OpenedWb = load_workbook(full_path )
    OpenedWS = OpenedWb['Sheet1']
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 4
    chart1.title = ""
    chart1.y_axis.title = ''
    chart1.x_axis.title = ''

    data = Reference(OpenedWS, min_col=2, min_row=1, max_row=ColumnCount)
    cats = Reference(OpenedWS, min_col=1, min_row=1, max_row=ColumnCount)
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.shape = 4
    OpenedWS.add_chart(chart1, "D3")
    OpenedWb.save(full_path)

def export_image(path):
    app = win32.Dispatch('Excel.Application')
    workbook_file_name = str(path) +  '\\source.xlsx'
    workbook = app.Workbooks.Open(Filename=workbook_file_name)
    app.DisplayAlerts = False

    for sheet in workbook.Worksheets:
        for chartObject in sheet.ChartObjects():
            chartObject.Activate()
            image_file_name=str(path) +  '\\image.png'
            chartObject.Chart.Export(image_file_name)
    workbook.Close(SaveChanges=False, Filename=workbook_file_name)


def generate(args):
    dirpath = os.path.join(args.BasePath, args.BaseName + '\\Data')
    if not os.access(str(dirpath), os.F_OK):
        os.makedirs(str(dirpath))
    for iteration in range(args.ChartsCount):
        iteration_path = os.path.join(dirpath, str(iteration+1))
        if not os.access(iteration_path, os.F_OK):
            os.makedirs(iteration_path)
        create_table(iteration_path)
        export_image(iteration_path)



def parse(argv):
    parser = argparse.ArgumentParser(description="""
     Parser of image and tables generator
        """, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('ChartsCount', action='store', type=int)
    parser.add_argument('BasePath', action='store', type=pathlib.Path)
    parser.add_argument('BaseName', action='store', type=str)

    return parser.parse_args(argv[1:])


def main():
    args = parse(sys.argv)
    generate(args)


main()
